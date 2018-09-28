#!/usr/bin/env python  
#coding:utf-8  
"""
@author:leo
@license: Apache Licence 
@file: analyseMain1.py
@time: 2018/08/14
@contact: leiaoo@163.com
"""
import os, datetime, re, json
import openpyxl

def main(flag):
    # currTimeStr = datetime.datetime.now().strftime('%Y-%m-%d')
    currTimeStr = '2018-08-09'
    baseDir = 'device'

    #获取文件名列表
    files = os.listdir(os.path.join(baseDir, flag, currTimeStr))
    #调用CreateReport函数，进行巡检文件分析及生成巡检报告
    #传入参数:
    #flag: 设备类型，Huawei、IOS、Nexus
    #currTimeStr: 以当前日期为名的文件夹
    #baseDir: 放置巡检文件的父文件夹，device
    #files: 获取到的文件名列表，list类型
    createReport(flag, currTimeStr, files, baseDir)

def createReport(flag, currTimeStr, files, baseDir):
    '''
    此函数用途：
    1、检查当次的巡检报告是否存在
    2、如果巡检报告不存在，创建新的巡检报告
    3、根据flag标志调用相应的巡检文件分析函数
    4、根据巡检文件分析函数返回的valuesDict字典，将所需的字段写入相应的表格中
    :param flag: 设备类型，Huawei、IOS、Nexus， str类型
    :param currTimeStr: 以当前日期为名的文件夹，str类型
    :param files: 获取到的文件名列表，list类型
    :param baseDir:  放置巡检文件的父文件夹，device， str类型
    :param valuesDict: 巡检结果文件分析后所需存储的字段，以字典方式保存， dict类型
    :return: None，不返回结果，只创建巡检报告文件
    '''
    #判断当次巡检报告文件是否存在，如果不存在，新建巡检报告文件
    if not os.path.exists(f'{currTimeStr}-report.xlsx'):
        wb = openpyxl.Workbook()
        wb.save(f'{currTimeStr}-report.xlsx')
    #加载巡检报告文件，获取对应的sheet，并以flag对应的标志命名
    wb = openpyxl.load_workbook(f'{currTimeStr}-report.xlsx')
    if 'Sheet' in wb.sheetnames:
        ws = wb['Sheet']
        ws.title = flag
    elif flag in wb.sheetnames:
        ws = wb[flag]
    else:
        ws = wb.create_sheet(title=flag)
    #设置表头
    ws['A1'] = "IP地址"
    ws['B1'] = '设备型号'
    ws['C1'] = '设备版本'
    ws['D1'] = '设备运行时间'
    ws['E1'] = 'CPU使用率'
    ws['F1'] = '内存使用率'
    ws['G1'] = 'Fex状态'
    ws['H1'] = '风扇状态'
    ws['I1'] = '电源状态'
    ws['J1'] = '温度状态'
    ws['K1'] = '百兆接口'
    ws['L1'] = 'errdisable接口'
    ws['M1'] = '设备健康状态'
    ws['N1'] = '设备状态'
    ws['O1'] = '根桥mac地址'
    ws['P1'] = '上一次重启时间'
    #设置变量，用于循环每行的数据
    wsline = 1
    for file in files:
        # define values Dict, default value is None
        valuesDict = {
            'IP': 'None',  # device file IP
            'flag': flag,  # device type
            'Version': 'None',  # device version
            'Type': 'None',  # device model
            'Bgp_Peer': 'None',  # bgp peer information
            'Ospf_Peer': 'None',  # ospf peer information
            'Ospf_brief': 'None',  # ospf information
            'Uptime': 'None',  # device up time
            'Health': 'None',  # device health information
            'Fan': 'None',  # device fan status
            'Power': 'None',  # device power status
            'Temperature': 'None',  # device temperature status
            'DeviceStatus': 'None',  # device status
            'CpuUsage': 'None',  # cpu usage
            'MemoryUsage': 'None',  # memory usage
            'ErrDisable': 'None',  # interface Error Disable reasone
            'treeRootMacadd': 'None',  # Spanning Tree Root address
            '100M': 'None',  # 100M interface
            'Fex': 'None',  # Fex status
            'RestartTime': 'None',  # device restarttime
            'InterfaceStatus': {  # interface information
                'InterfaceNum': 'None',  # interface Number
                'PHY': 'None',  # interface physics status
                'Protocol': 'None',  # interface protocol status
                'InputErrors': 'None',  # interface input error
                'OutputErrors': 'None',  # interface output error
                'InUti': 'None',  # interface input utilities
                'OutUti': 'None',  # interface output utilities
                'Speed': 'None',  # interface speed
                'Duplex': 'None',  # interface duplex module
            },
        }
        wsline += 1 #从第2行开始写入数据
        if flag == 'Nexus': #调用分析Nexus巡检文件的函数
            valuesDict = analyseNexus(flag, file, currTimeStr, baseDir, valuesDict)
        elif flag == 'IOS': #调用分析IOS巡检文件的函数
            valuesDict = analyseIOS(flag, file, currTimeStr, baseDir, valuesDict)
        else: #调用分析Huawei巡检文件的函数
            valuesDict = analyseHuawei(flag, file, currTimeStr, baseDir, valuesDict)
        print(valuesDict['treeRootMacadd'])
        ws[f'A{wsline}'] = valuesDict['IP']
        ws[f'B{wsline}'] = valuesDict['Type']
        ws[f'C{wsline}'] = valuesDict['Version']
        ws[f'D{wsline}'] = valuesDict['Uptime']
        ws[f'E{wsline}'] = valuesDict['CpuUsage']
        ws[f'F{wsline}'] = valuesDict['MemoryUsage']
        ws[f'G{wsline}'] = valuesDict['Fex']
        ws[f'H{wsline}'] = valuesDict['Fan']
        ws[f'I{wsline}'] = valuesDict['Power']
        ws[f'J{wsline}'] = valuesDict['Temperature']
        ws[f'K{wsline}'] = valuesDict['100M']
        ws[f'L{wsline}'] = valuesDict['ErrDisable']
        ws[f'M{wsline}'] = valuesDict['Health']
        ws[f'N{wsline}'] = valuesDict['DeviceStatus']
        ws[f'O{wsline}'] = valuesDict['treeRootMacadd']
        ws[f'P{wsline}'] = valuesDict['RestartTime']

    wb.save(f'{currTimeStr}-report.xlsx')

def analyseNexus(flag, file, currTimeStr, baseDir, valuesDict):
    '''
    此函数用途：
    1、分析Nexus设备巡检文件
    2、获取需要判断和监控的字段
    :param flag: 设备类型，Huawei、IOS、Nexus，类型：str
    :param file: 巡检文件，类型：str
    :param currTimeStr: 当次目录，类型：str
    :param baseDir: 巡检文件的父目录，类型：str
    :param valuesDict: 用于存储巡检结果字段，类型：dict
    :return:  valuesDict，类型：dict
    '''
    ipStrRe = re.compile(r'(\d{1,3}(\.)?){1,4}')  # get ip address re
    cpuStrRe = re.compile(r'\d{0,2}\.?\d{0,2}%')  # get cpu usage rate
    memoryRe = re.compile(r'\d{1,}')  # get memory usage rate
    # get command list
    with open(f'conf_files/NexusCommand.txt', 'r') as f:
        CommandString = json.load(f)
    #定义一个空字典，用于获取每个命令结果，以列表方式存储
    valuesList = {}
    #获取当前分析文件名中的IP地址
    valuesDict['IP'] = ipStrRe.search(file).group()
    #获取当前操作文件的绝对路径
    file = os.path.join(baseDir, flag, currTimeStr, file)
    #打开当前要分析的文件
    with open(file, 'r') as f:
        fileReadLines = f.readlines()
    #获取每个命令的结果，以字典-列表方式存储
    for line in range(1, len(CommandString)):
        if line == len(CommandString) - 1:
            valuesList[CommandString[line]] = fileReadLines[fileReadLines.index(CommandString[line] + '\n'):]
        valuesList[CommandString[line - 1]] = fileReadLines[fileReadLines.index(CommandString[line - 1] + '\n'):
                                                             fileReadLines.index(CommandString[line] + '\n')]
    #获取所需的字段
    for item in CommandString[1:]:
        #获取设备版本、型号、运行时间
        if item == 'show version':
            valuesDict['Version'] = [line.strip(' ').split(' ')[-1].strip('\n') for line in valuesList[item]
                               if 'system version:' in line.lower()][0]
            valuesDict['Type'] = [line.strip(' ').split(' ')[2] for line in valuesList[item]
                               if 'cisco nexus' in line.lower() and 'chassis' in line.lower()][0]
            valuesDict['Uptime'] = ' '.join([line.split(' ')[3:5] for line in valuesList[item]
                                        if 'uptime is' in line.lower()][0]).strip('(s),')
        #获取设备风扇、电源、温度
        if item == 'show env':
            for line in valuesList[item][valuesList[item].index('Fan:\n') + 4:]:
                if line == '\n':
                    break
                elif ':' in line:
                    break
                elif 'ok' not in line.lower():
                    valuesDict['Fan'] = 'Failed'
                    break
                else:
                    valuesDict['Fan'] = 'Ok'
            for line in valuesList[item][valuesList[item].index('Power Supply:\n') + 6:]:
                if line == '\n':
                    break
                elif 'ok' not in line.lower():
                    valuesDict['Power'] = 'Faild'
                    break
                else:
                    valuesDict['Power'] = 'Ok'
            temperatureIndex = None
            for line in valuesList[item]:
                if 'Temperature' in line:
                    temperatureIndex = valuesList[item].index(line)
            for line in valuesList[item][temperatureIndex + 5:]:
                if line == '\n':
                    break
                if 'ok' not in line.lower():
                    valuesDict['Temperature'] = 'Failed'
                    break
                else:
                    valuesDict['Temperature'] = 'Ok'
        #获取fex设备的风扇、温度、电源信息
        if item == 'show env fex all':
            if len(valuesList[item]) > 3:
                for line in valuesList[item][1:]:
                    if ' Fex ' in line:
                        FexNum = line.split(' ')[-1].strip(':\n')
                        if f'Fan Fex: {FexNum}:\n' == line:
                            for xline in valuesList[item][valuesList[item].index(f'Fan Fex: {FexNum}:\n') + 4:]:
                                if xline == '\n':
                                    break
                                elif 'ok' not in xline.lower():
                                    valuesDict[f'Fex{FexNum}Fan'] = 'Failed'
                                else:
                                    valuesDict[f'Fex'] = 'Ok'
                        elif f'Temperature Fex {FexNum}:\n' == line:
                            for xline in valuesList[item][valuesList[item].index(f'Temperature Fex {FexNum}:\n') + 5:]:
                                if xline == '\n':
                                    break
                                elif 'ok' not in xline.lower():
                                    valuesDict[f'Fex{FexNum}Temperature'] = 'Failed'
                                else:
                                    valuesDict[f'Fex'] = 'Ok'
                        elif f'Power Supply Fex {FexNum}:\n' == item:
                            for xline in valuesList[item][
                                        valuesList[item].index(f'Power Supply Fex {FexNum}:\n') + 7:]:
                                if xline == '\n':
                                    break
                                elif 'ok' not in xline.lower():
                                    valuesDict[f'Fex{FexNum}Power'] = 'Failed'
                                else:
                                    valuesDict[f'Fex'] = 'Ok'
                        else:
                            continue
        #获取设备module信息
        if item == 'show module':
            for line in valuesList[item][3:]:
                if line == '\n':
                    break
                if 'active' not in line:
                    if 'ok' not in line:
                        if 'standby' not in line:
                            valuesDict['Module'] = 'Failed'
                            break
                else:
                    valuesDict['Module'] = 'Ok'
        #获取cpu、内存信息
        if item == 'show system resource':
            for line in valuesList[item]:
                if 'CPU states' in line:
                    valuesDict['CpuUsage'] = cpuStrRe.search(valuesList[item][3]).group()
                if 'Memory usage:' in line:
                    valuesDict['MemoryUsage'] = '{:.2f}%'.format(int(memoryRe.findall(line)[1]) /
                                                                 int(memoryRe.findall(line)[0]) * 100)
        #获取100M接口信息
        if item == 'show inter status | in "a-100 "':
            if len(valuesList[item]) > 3:
                valuesDict['100M'] = ','.join([line.split(' ')[0] for line in valuesList[item][1:]
                                                if 'a-100' in line.lower()])
        #获取error-disable接口信息
        if item == 'show inter status err-dis':
            if valuesList[item][-1] == '\n' or '-' * 5 in valuesList[item][-1]:
                pass
            else:
                valuesDict['ErrDisable'] = ''.join([line.split(' ')[0] for line in valuesList[item][5:]])
        #获取设备根桥信息
        if item == 'show spanning-tree root' and len(valuesList[item]) > 3:
            treeRootReg = re.compile(r'(\w{4}(\.)?){3}')
            treeRootMacadd = []
            for line in valuesList[item]:
                if treeRootReg.search(line) == None:
                    continue
                treeRootMacadd += [treeRootReg.search(line).group()]
            valuesDict['treeRootMacadd'] = '\n'.join(set(treeRootMacadd))
    return valuesDict

def analyseIOS(flag, file, currTimeStr, baseDir, valuesDict):
    '''
    此函数用途：
    1、分析IOS设备巡检文件
    2、获取需要判断和监控的字段
    :param flag: 设备类型，Huawei、IOS、Nexus，类型：str
    :param file: 巡检文件，类型：str
    :param currTimeStr: 当次目录，类型：str
    :param baseDir: 巡检文件的父目录，类型：str
    :param valuesDict: 用于存储巡检结果字段，类型：dict
    :return:  valuesDict，类型：dict
    '''
    ipStrRe = re.compile(r'(\d{1,3}(\.)?){1,4}')  # get ip address re
    cpuStrRe = re.compile(r'\d{0,2}\.?\d{0,2}%')  # get cpu usage rate
    memoryRe = re.compile(r'\d{1,}')  # get memory usage rate
    # get command list
    with open(f'conf_files/IosCommand.txt', 'r') as f:
        CommandString = json.load(f)
    #定义一个空字典，用于获取每个命令结果，以列表方式存储
    valuesList = {}
    #获取当前分析文件名中的IP地址
    valuesDict['IP'] = ipStrRe.search(file).group()
    #获取当前操作文件的绝对路径
    file = os.path.join(baseDir, flag, currTimeStr, file)
    #打开当前要分析的文件
    with open(file, 'r') as f:
        fileReadLines = f.readlines()
    #获取每个命令的结果，以字典-列表方式存储
    for line in range(1, len(CommandString)):
        if line == len(CommandString) - 1:
            valuesList[CommandString[line]] = fileReadLines[fileReadLines.index(CommandString[line] + '\n'):]
        valuesList[CommandString[line - 1]] = fileReadLines[fileReadLines.index(CommandString[line - 1] + '\n'):
                                                             fileReadLines.index(CommandString[line] + '\n')]
    #获取所需的字段
    for item in CommandString[1:]:
        #获取设备版本、型号、运行时间
        if item == 'show version':
            valuesDict['Version'] = [line.split(',')[2].strip(' ').split(' ')[1] for line in valuesList[item]
                                     if 'cisco ios software' in line.lower() and 'software' in line.lower()][0]
            valuesDict['Type'] = [line.split(',')[1].strip(' ').split(' ')[0] for line in valuesList[item]
                                  if 'cisco ios software' in line.lower() and 'software' in line.lower()][0]
            valuesDict['Uptime'] = ''.join([line.split(' ') for line in valuesList[item]
                                            if 'uptime is' in line.lower()][0][3:9]).strip(',')
            restartTime = [line.split(' ')[3:] for line in valuesList[item]
                           if 'system restarted' in line.lower()][0]
            restartTime.pop(1)
            valuesDict['RestartTime'] = '-'.join(restartTime).strip('\n')
        #获取设备风扇、电源、温度
        if item == 'show env all':
            fan = [line.strip('\n') for line in valuesList[item]
                   if 'fan' in line.lower()]
            for line in fan:
                if 'normal' in line.lower() or 'fan speed' in line.lower() or 'ok' in line.lower():
                    valuesDict['Fan'] = 'OK'
                else:
                    valuesDict['Fan'] = 'Failed'
                    break
            temperature = ', '.join([line.strip('\n') for line in valuesList[item]
                                     if 'system temperature' in line.lower()])
            if 'ok' in temperature.lower():
                valuesDict['Temperature'] = "OK"
            elif temperature == '':
                pass
            else:
                valuesDict['Temperature'] = "failed"

            power = [line.strip('\n') for line in valuesList[item]
                     if 'power supply' in line.lower() or 'temp: pem' in line.lower()]
            for line in power:
                if 'normal' in line.lower() or 'ok' in line.lower():
                    valuesDict['Power'] = 'OK'
                else:
                    valuesDict['Power'] = 'Failed'
                    break
        #获取cpu信息
        if item == 'show process cpu':
            valuesDict['CpuUsage'] = ' '.join([line.split(' ')[-3:] for line in valuesList[item]
                                if 'cpu utilization' in line.lower()][0]).split(':')[1].strip(' ').strip('\n')
        # 获取内存信息
        if item == 'show proc mem':
            TotalMemory = [line.split(':')[1].strip(' ').split(' ')[0] for line in valuesList[item]
                           if 'processor pool' in line.lower()]
            UsedMemory = [line.split(':')[2].strip(' ').split(' ')[0] for line in valuesList[item]
                          if 'processor pool' in line.lower()]
            valuesDict['MemoryUsage'] = '{:.2f}%'.format((int(UsedMemory[0]) / int(TotalMemory[0]) * 100))
        #获取100M接口信息
        if item == 'show inter status | i (a-100 )':
            a_100 = ','.join([line.split(' ')[0] for line in valuesList[item][1:]
                              if 'a-100' in line.lower()])
            if a_100 == '':
                pass
            else:
                valuesDict['100M'] = a_100
        #获取error-disable接口信息
        if item == 'show inter status err-disabled':
            errdisable = [line.split(' ')[0] for line in valuesList[item][1:]
                          if 'err' in line.lower()]
            if len(errdisable) > 1:
                valuesDict['ErrDisable'] = errdisable[1]
        # 获取设备根桥信息
        if item == 'show spanning-tree root':
            treeRoot = [line for line in valuesList[item][4:]]
            if len(treeRoot) > 1:
                treeRootReg = re.compile(r'(\w{4}(\.)?){3}')
                treeRootMacadd = ''
                for line in treeRoot[1:]:
                    treeRootMacadd = treeRootReg.search(line).group()
                valuesDict['treeRootMacadd'] = treeRootMacadd
    return valuesDict

def analyseHuawei(flag, file, currTimeStr, baseDir, valuesDict):
    '''
    此函数用途：
    1、分析Huawei设备巡检文件
    2、获取需要判断和监控的字段
    :param flag: 设备类型，Huawei、IOS、Nexus，类型：str
    :param file: 巡检文件，类型：str
    :param currTimeStr: 当次目录，类型：str
    :param baseDir: 巡检文件的父目录，类型：str
    :param valuesDict: 用于存储巡检结果字段，类型：dict
    :return:  valuesDict，类型：dict
    '''
    ipStrRe = re.compile(r'(\d{1,3}(\.)?){1,4}')  # get ip address re
    cpuStrRe = re.compile(r'\d{0,2}\.?\d{0,2}%')  # get cpu usage rate
    memoryRe = re.compile(r'\d{1,}')  # get memory usage rate
    # get command list
    with open(f'conf_files/HuaweiCommand.txt', 'r') as f:
        CommandString = json.load(f)
    #定义一个空字典，用于获取每个命令结果，以列表方式存储
    valuesList = {}
    #获取当前分析文件名中的IP地址
    valuesDict['IP'] = ipStrRe.search(file).group()
    #获取当前操作文件的绝对路径
    file = os.path.join(baseDir, flag, currTimeStr, file)
    #打开当前要分析的文件
    with open(file, 'r') as f:
        fileReadLines = f.readlines()
    #获取每个命令的结果，以字典-列表方式存储
    for line in range(1, len(CommandString)):
        if line == len(CommandString) - 1:
            valuesList[CommandString[line]] = fileReadLines[fileReadLines.index(CommandString[line] + '\n'):]
        valuesList[CommandString[line - 1]] = fileReadLines[fileReadLines.index(CommandString[line - 1] + '\n'):
                                                             fileReadLines.index(CommandString[line] + '\n')]
    #获取所需的字段
    for item in CommandString:
        #获取设备版本、型号、运行时间
        if item == 'display version':
            version = [line.split(',')[1].strip('\n') for line in valuesList[item]
                       if 'software, version' in line.lower()][0].split(' ')
            valuesDict['Version'] = version[2]
            valuesDict['Type'] = version[3].strip('(')
            valuesDict['Uptime'] = str(int([line.split(':')[1].strip('\n') for line in valuesList[item]
                            if ': uptime' in line.lower()][0].strip(' ').split(' ')[2]) * 7)
        #获取设备风扇
        if item == 'display fan':
            if len(valuesList[item]) > 3:
                valuesDict['Fan'] = 'Normal' if 'normal' in valuesList[item][4].lower() or 'yes' in \
                                            valuesList[item][4].lower() else 'Failed'
        # 获取设备电源
        if item == 'display power':
            if len(valuesList[item]) > 3:
                failedFlag = 'Normal'
                for line in valuesList[item][4:-1]:
                    if '-' * 10 in line or  line == '\n':
                        break
                    elif 'normal' in line.lower() and 'yes' in line.lower():
                        continue
                    elif 'ac' in line.lower() and 'supply' in line.lower():
                        continue
                    else:
                        failedFlag = 'Failed'
                        break
                valuesDict['Power'] = failedFlag
        # 获取设备温度
        if item == 'display temperature' or item == 'display temperature all':
            if len(valuesList[item]) > 3:
                failedFlag = 'Normal'
                for line in valuesList[item][4:-1]:
                    if '-' * 10 in line or line == '\n':
                        break
                    elif 'normal' in line.lower():
                        continue
                    elif 'yes' in line.lower():
                        continue
                    else:
                        failedFlag = 'Failed'
                        break
                valuesDict['Temperature'] = failedFlag
        # 获取设备健康状态
        if item == 'display health':
            if len(valuesList[item]) > 3:
                failedFlag = 'Normal'
                for line in valuesList[item][4:9]:
                    if '-' * 10 in line or line == '\n':
                        break
                    elif 'normal' in line.lower():
                        continue
                    elif 'yes' in line.lower():
                        continue
                    else:
                        failedFlag = 'Failed'
                        break
                valuesDict['Health'] = failedFlag
        # 获取设备状态
        if item == 'display device':
            if len(valuesList[item]) > 3:
                failedFlag = 'Normal'
                for line in valuesList[item][4:-1]:
                    if '-' * 10 in line or line == '\n':
                        break
                    elif 'normal' in line.lower():
                        continue
                    elif 'yes' in line.lower():
                        continue
                    else:
                        failedFlag = 'Failed'
                        break
                valuesDict['DeviceStatus'] = failedFlag
        #获取cpu信息
        if item == 'display cpu-usage':
            cpuStr = [line for line in valuesList[item] if 'cpu usage'
                      in line.lower() and 'max' in line.lower()][0].strip('\n').strip(' ')
            valuesDict['CpuUsage'] = cpuStrRe.search(cpuStr).group()
        # 获取内存信息
        if item == 'display memory-usage':
            valuesDict['MemoryUsage'] = [line.split(':')[1].strip('\n') for line in valuesList[item]
                                         if 'memory using' in line.lower()][0].strip(' ')
    return valuesDict

main('Nexus')
# main('IOS')
# main('Huawei')
