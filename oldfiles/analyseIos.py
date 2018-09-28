#!/usr/bin/env python
#coding:utf-8  
"""
@author:leo
@license: Apache Licence 
@file: analyseIos.py 
@time: 2018/08/08
@contact: leiaoo@163.com
"""
import os, datetime, re, json
import openpyxl

ipStrRe = re.compile(r'(\d{1,3}(\.)?){1,4}')
cpuStrRe = re.compile(r'\d{0,2}\.?\d{0,2}%')
# currTimeStr = datetime.datetime.now().strftime('%Y-%m-%d')
currTimeStr = '2018-08-09'
baseDir = 'device'
flag = 'IOS'
files = os.listdir(os.path.join(baseDir, flag, currTimeStr))
with open('conf_files/IosCommand.txt', 'r') as f:
    CommandString = json.load(f)
valuesList = {}

# def processList(valuesList, startIndex, endIndex):
#     failedFlag = 'Normal'
#     for line in valuesList[startIndex:endIndex]:
#         if '-' * 10 in line:
#             continue
#         elif line == '\n':
#             continue
#         elif 'normal' in line.lower():
#             continue
#         elif 'yes' in line.lower():
#             continue
#         elif 'ac' in line.lower() and 'supply' in line.lower():
#             continue
#         else:
#             failedFlag = 'Failed'
#             break
#     return failedFlag

def processHuaWeiFile(file, CommandString, ):
    valuesDict = {
        'IP': 'None', #device file IP
        'flag': flag, #device type
        'Version': 'None', #device version
        'UpTime': 'None',  #device up time
        'Health': 'None',  #device health information
        'Fan': 'None',  #device fan status
        'Power': 'None', #device power status
        'Temperature': 'None', #device temperature status
        'DeviceStatus': 'None', #device status
        'CpuUsage': 'None', #cpu usage
        'MemoryUsage': 'None', # memory usage
        'ErrDisable' : 'None',
        'treeRootMacadd' : 'None',
        '100M' : 'None',
        'InterfaceStatus': {  #interface information
            'InterfaceNum': 'None', #interface Number
            'PHY': 'None',  #interface physics status
            'Protocol': 'None', #interface protocol status
            'InputErrors': 'None', #interface input error
            'OutputErrors': 'None', #interface output error
            'InUti': 'None', #interface input utilities
            'OutUti': 'None', #interface output utilities
            'Speed': 'None', #interface speed
            'Duplex': 'None', #interface duplex module
        },
    }
    commandIndex = {}
    valuesDict['IP'] = ipStrRe.search(file).group()
    file = os.path.join(baseDir, flag, currTimeStr, file)

    with open(file, 'r') as f:
        fileReadLines = f.readlines()

    for i in range(len(CommandString)):
        commandIndex[CommandString[i]] = fileReadLines.index(CommandString[i] + '\n')

    valuesList['version'] = fileReadLines[commandIndex['show version']:
                                          commandIndex['show env all']]
    valuesList['envAll'] = fileReadLines[commandIndex['show env all']:
                                         commandIndex['show module']]
    valuesList['module'] = fileReadLines[commandIndex['show module']:
                                         commandIndex['show process cpu']]
    valuesList['cpu'] = fileReadLines[commandIndex['show process cpu']:
                                      commandIndex['show process cpu history']]
    valuesList['cpuHistory'] = fileReadLines[commandIndex['show process cpu history']:
                                             commandIndex['show proc mem']]
    valuesList['memory'] = fileReadLines[commandIndex['show proc mem']:
                                         commandIndex['show interface']]
    valuesList['showInt'] = fileReadLines[commandIndex['show interface']:
                                          commandIndex['show inter status']]
    valuesList['IntStatus'] = fileReadLines[commandIndex['show inter status']:
                                            commandIndex['show inter status | i (a-100 )']]
    valuesList['100M'] = fileReadLines[commandIndex['show inter status | i (a-100 )']:
                                       commandIndex['show inter status err-disabled']]
    valuesList['errdisable'] = fileReadLines[commandIndex['show inter status err-disabled']:
                                             commandIndex['show hsrp brief']]
    valuesList['hsrp'] = fileReadLines[commandIndex['show hsrp brief']:
                                       commandIndex['show spanning-tree root']]
    valuesList['spanningtree'] = fileReadLines[commandIndex['show spanning-tree root']:
                                               commandIndex['show port-channel sum']]
    valuesList['port-channle'] = fileReadLines[commandIndex['show port-channel sum']:
                                               commandIndex['show logging']]
    valuesList['logging'] = fileReadLines[commandIndex['show logging']:]

    valuesDict['Version'] = [line.split(',')[2].strip(' ').split(' ')[1] for line in valuesList['version']
                            if 'Cisco IOS Software'.lower() in line.lower() and 'software'.lower() in line.lower()][0]
    valuesDict['Type'] = [line.split(',')[1].strip(' ').split(' ')[0] for line in valuesList['version']
                            if 'Cisco IOS Software'.lower() in line.lower() and 'software'.lower() in line.lower()][0]
    valuesDict['UpTime'] = ''.join([line.split(' ') for line in valuesList['version']
                             if 'uptime is'.lower() in line.lower()][0][3:9]).strip(',')
    restartTime = [line.split(' ')[3:] for line in valuesList['version']
                                          if 'System restarted'.lower() in line.lower()][0]
    restartTime.pop(1)
    valuesDict['RestartTime'] = '-'.join(restartTime).strip('\n')

    fan = [line.strip('\n') for line in valuesList['envAll']
                                          if 'fan' in line.lower()]
    for item in fan:
        if 'normal' in item.lower() or 'fan speed' in item.lower() or 'ok' in item.lower():
            valuesDict['Fan'] = 'OK'
        else:
            valuesDict['Fan'] = 'Failed'
            break
    temperature = ', '.join([line.strip('\n') for line in valuesList['envAll']
                                   if 'SYSTEM TEMPERATURE'.lower() in line.lower()])
    if 'ok' in temperature.lower():
        valuesDict['Temperature'] = "OK"
    elif temperature == '':
        pass
    else:
        valuesDict['Temperature'] = "failed"

    power = [line.strip('\n') for line in valuesList['envAll']
                                           if 'POWER SUPPLY'.lower() in line.lower() or 'temp: pem' in line.lower()]
    for item in power:
        if 'normal' in item.lower() or 'ok' in item.lower():
            valuesDict['Power'] = 'OK'
        else:
            valuesDict['Power'] = 'Failed'
            break
    valuesDict['CpuUsage'] = ' '.join([line.split(' ')[-3:] for line in valuesList['cpu']
                            if 'CPU utilization'.lower() in line.lower()][0]).split(':')[1].strip(' ').strip('\n')
    TotalMemory = [line.split(':')[1].strip(' ').split(' ')[0] for line in valuesList['memory']
                                       if 'Processor Pool'.lower() in line.lower()]
    UsedMemory = [line.split(':')[2].strip(' ').split(' ')[0] for line in valuesList['memory']
                                       if 'Processor Pool'.lower() in line.lower()]
    valuesDict['MemoryUsage'] = '{:.2f}%'.format((int(UsedMemory[0]) / int(TotalMemory[0]) * 100))

    a_100 = ','.join([line.split(' ')[0] for line in valuesList['100M'][1:]
                          if 'a-100' in line.lower()])
    if a_100 == '':
        pass
    else:
        valuesDict['100M'] = a_100
    errdisable = [line.split(' ')[0] for line in valuesList['errdisable'][1:]
                                        if 'err' in line.lower()]
    if len(errdisable) > 1:
        valuesDict['ErrDisable'] = errdisable[1]

    treeRoot = [line for line in valuesList['spanningtree'][4:]]
    if len(treeRoot) > 1:
        treeRootReg = re.compile(r'(\w{4}(\.)?){3}')
        treeRootMacadd = ''
        for item in treeRoot[1:]:
            treeRootMacadd = treeRootReg.search(item).group()
        valuesDict['treeRootMacadd'] = treeRootMacadd

    return valuesDict

if not os.path.exists(f'{currTimeStr}-report.xlsx'):
    wb = openpyxl.Workbook()
    wb.save(f'{currTimeStr}-report.xlsx')
wb = openpyxl.load_workbook(f'{currTimeStr}-report.xlsx')
if flag in wb.sheetnames:
    ws_ios = wb[flag]
else:
    ws_ios = wb.create_sheet(title=flag)

ws_ios['A1'] = "IP地址"
ws_ios['B1'] = '设备型号'
ws_ios['C1'] = '设备版本'
ws_ios['D1'] = '设备运行时间'
ws_ios['E1'] = '设备重启时间'
ws_ios['F1'] = 'CPU使用率'
ws_ios['G1'] = '内存使用率'
ws_ios['H1'] = '风扇状态'
ws_ios['I1'] = '电源状态'
ws_ios['J1'] = '温度状态'
ws_ios['K1'] = '百兆接口'
ws_ios['L1'] = 'errdisable接口'
ws_ios['M1'] = '根桥地址'

wsline = 1
for file in files:
    wsline += 1
    valuesDict = processHuaWeiFile(file, CommandString)
    # print(valuesDict['RestartTime'])
    ws_ios[f'A{wsline}'] = valuesDict['IP']
    ws_ios[f'B{wsline}'] = valuesDict['Type']
    ws_ios[f'C{wsline}'] = valuesDict['Version']
    ws_ios[f'D{wsline}'] = valuesDict['UpTime']
    ws_ios[f'E{wsline}'] = valuesDict['RestartTime']
    ws_ios[f'F{wsline}'] = valuesDict['CpuUsage']
    ws_ios[f'G{wsline}'] = valuesDict['MemoryUsage']
    ws_ios[f'H{wsline}'] = valuesDict['Fan']
    ws_ios[f'I{wsline}'] = valuesDict['Power']
    ws_ios[f'J{wsline}'] = valuesDict['Temperature']
    ws_ios[f'K{wsline}'] = valuesDict['100M']
    ws_ios[f'L{wsline}'] = valuesDict['ErrDisable']
    ws_ios[f'M{wsline}'] = valuesDict['treeRootMacadd']
wb.save(f'{currTimeStr}-report.xlsx')
