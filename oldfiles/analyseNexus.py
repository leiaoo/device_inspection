#!/usr/bin/env python  
#coding:utf-8  
"""
@author:leo
@license: Apache Licence 
@file: analyseNexus.py
@time: 2018/08/09
@contact: leiaoo@163.com
"""
import os, datetime, re, json, pprint
import openpyxl

ipStrRe = re.compile(r'(\d{1,3}(\.)?){1,4}')
cpuStrRe = re.compile(r'\d{0,2}\.?\d{0,2}%')
memoryRe = re.compile(r'\d{1,}')
# currTimeStr = datetime.datetime.now().strftime('%Y-%m-%d')
currTimeStr = '2018-08-09'
baseDir = 'device'
flag = 'Nexus'
files = os.listdir(os.path.join(baseDir, flag, currTimeStr))
with open('conf_files/NexusCommand.txt', 'r') as f:
    CommandString = json.load(f)
def processHuaWeiFile(file, CommandString, ):
    valuesDict = {
        'IP': 'None', #device file IP
        'flag': flag, #device type
        'Version': 'None', #device version
        'UpTime': 'None',  #device up time
        'Health': 'None',  #device health information
        'Fan': 'None',  #device fan status
        'Power': 'None', #device power status
        'Temperature': 'None', #device temperature status
        'DeviceStatus': 'None', #device status
        'CpuUsage': 'None', #cpu usage
        'MemoryUsage': 'None', # memory usage
        'ErrDisable' : 'None',
        'treeRootMacadd' : 'None',
        '100M' : 'None',
        'Fex' : 'None',
        'InterfaceStatus': {  #interface information
            'InterfaceNum': 'None', #interface Number
            'PHY': 'None',  #interface physics status
            'Protocol': 'None', #interface protocol status
            'InputErrors': 'None', #interface input error
            'OutputErrors': 'None', #interface output error
            'InUti': 'None', #interface input utilities
            'OutUti': 'None', #interface output utilities
            'Speed': 'None', #interface speed
            'Duplex': 'None', #interface duplex module
        },
    }
    commandIndex = {}
    valuesList = {}
    valuesList1 = {}
    valuesDict['IP'] = ipStrRe.search(file).group()
    file = os.path.join(baseDir, flag, currTimeStr, file)
    with open(file, 'r') as f:
        fileReadLines = f.readlines()

    for i in range(len(CommandString)):
        commandIndex[CommandString[i]] = fileReadLines.index(CommandString[i] + '\n')

    valuesList['version'] = fileReadLines[commandIndex['show version']:
                                          commandIndex['show env']]
    valuesList['env'] = fileReadLines[commandIndex['show env']:
                                         commandIndex['show env fex all ']]
    valuesList['fex'] = fileReadLines[commandIndex['show env fex all ']:
                                         commandIndex['show module']]
    valuesList['module'] = fileReadLines[commandIndex['show module']:
                                      commandIndex['show processes cpu']]
    valuesList['cpu'] = fileReadLines[commandIndex['show processes cpu']:
                                      commandIndex['show processes cpu history']]
    valuesList['cpuHistory'] = fileReadLines[commandIndex['show processes cpu history']:
                                             commandIndex['show system resource ']]
    valuesList['resource'] = fileReadLines[commandIndex['show system resource ']:
                                         commandIndex['show interface']]
    valuesList['showInt'] = fileReadLines[commandIndex['show interface']:
                                          commandIndex['show inter statu']]
    valuesList['IntStatus'] = fileReadLines[commandIndex['show inter statu']:
                                            commandIndex['show inter status | in "a-100 "']]
    valuesList['100M'] = fileReadLines[commandIndex['show inter status | in "a-100 "']:
                                       commandIndex['show inter status err-dis']]
    valuesList['errdisable'] = fileReadLines[commandIndex['show inter status err-dis']:
                                             commandIndex['show vpc brief']]
    valuesList['vpc'] = fileReadLines[commandIndex['show vpc brief']:
                                             commandIndex['show hsrp brief']]
    valuesList['hsrp'] = fileReadLines[commandIndex['show hsrp brief']:
                                       commandIndex['show ip ospf nei']]
    valuesList['ospf'] = fileReadLines[commandIndex['show ip ospf nei']:
                                               commandIndex['show port-channel sum']]
    valuesList['port-channel'] = fileReadLines[commandIndex['show port-channel sum']:
                                               commandIndex['show logging']]
    valuesList['logging'] = fileReadLines[commandIndex['show logging']:]

    valuesDict['Version'] = [line.strip(' ').split(' ')[-1].strip('\n') for line in valuesList['version']
                            if 'system version:' in line.lower()][0]
    valuesDict['Type'] = [line.strip(' ').split(' ')[2] for line in valuesList['version']
                            if 'cisco nexus' in line.lower() and 'chassis' in line.lower()][0]
    valuesDict['UpTime'] = ' '.join([line.split(' ')[3:5] for line in valuesList['version']
                             if 'uptime is' in line.lower()][0]).strip('(s),')

    for line in valuesList['env'][valuesList['env'].index('Fan:\n') + 4 :]:
        if line == '\n':
            break
        elif ':' in line:
            break
        elif 'ok' not in line.lower():
            valuesDict['Fan'] = 'Failed'
        else:
            valuesDict['Fan'] = 'Ok'

    for line in valuesList['env'][valuesList['env'].index('Power Supply:\n') + 6 :]:
        if line == '\n':
            break
        elif 'ok' not in line.lower():
            valuesDict['Power'] = 'Faild'
            break
        else:
            valuesDict['Power'] = 'Ok'
    temperatureIndex = None
    for item in valuesList['env']:
        if 'Temperature' in item:
            temperatureIndex = valuesList['env'].index(item)
    for item in valuesList['env'][temperatureIndex + 5:]:
        if item == '\n':
            break
        if 'ok' not in item.lower():
            valuesDict['Temperature'] = 'Failed'
            break
        else:
            valuesDict['Temperature'] = 'Ok'
    if len(valuesList['fex']) > 3:
        for item in valuesList['fex'][1:]:
            if ' Fex ' in item:
                FexNum = item.split(' ')[-1].strip(':\n')
                if f'Fan Fex: {FexNum}:\n' == item:
                    for line in valuesList['fex'][valuesList['fex'].index(f'Fan Fex: {FexNum}:\n') + 4:]:
                        if line == '\n':
                            break
                        elif 'ok' not in line.lower():
                            valuesDict[f'Fex{FexNum}Fan'] = 'Failed'
                        else:
                            valuesDict[f'Fex'] = 'Ok'
                elif f'Temperature Fex {FexNum}:\n' == item:
                    for line in valuesList['fex'][valuesList['fex'].index(f'Temperature Fex {FexNum}:\n') + 5:]:
                        if line == '\n':
                            break
                        elif 'ok' not in line.lower():
                            valuesDict[f'Fex{FexNum}Temperature'] = 'Failed'
                        else:
                            valuesDict[f'Fex'] = 'Ok'
                elif f'Power Supply Fex {FexNum}:\n' == item:
                    for line in valuesList['fex'][valuesList['fex'].index(f'Power Supply Fex {FexNum}:\n') + 7:]:
                        if line == '\n':
                            break
                        elif 'ok' not in line.lower():
                            valuesDict[f'Fex{FexNum}Power'] = 'Failed'
                        else:
                            valuesDict[f'Fex'] = 'Ok'
                else:
                    continue

    for item in valuesList['module'][3:]:
        if item == '\n':
            break
        if 'active' not in item:
            if 'ok' not in item:
                if 'standby' not in item:
                    valuesDict['Module'] = 'Failed'
                    break
        else:
            valuesDict['Module'] = 'Ok'

    for item in valuesList['resource']:
        if 'CPU states' in item:
            valuesDict['CpuUsage'] = cpuStrRe.search(valuesList['resource'][3]).group()
        if 'Memory usage:' in item:
            valuesDict['MemoryUsage'] = '{:.2f}%'.format(int(memoryRe.findall(item)[1])/
                                                         int(memoryRe.findall(item)[0]) * 100)
    if len(valuesList['100M']) > 3:
        valuesDict['a-100'] = ','.join([line.split(' ')[0] for line in valuesList['100M'][1:]
                  if 'a-100' in line.lower()])
    if valuesList['errdisable'][-1] == '\n' or '-' * 5 in valuesList['errdisable'][-1]:
        pass
    else:
        valuesDict['ErrDisable'] =''.join([line.split(' ')[0] for line in valuesList['errdisable'][5:]])

    # treeRoot = [line for line in valuesList['spanningtree'][4:]]
    # if len(treeRoot) > 1:
    #     treeRootReg = re.compile(r'(\w{4}(\.)?){3}')
    #     treeRootMacadd = ''
    #     for item in treeRoot[1:]:
    #         treeRootMacadd = treeRootReg.search(item).group()
    #     valuesDict['treeRootMacadd'] = treeRootMacadd

    return valuesDict

if not os.path.exists(f'{currTimeStr}-report.xlsx'):
    wb = openpyxl.Workbook()
    wb.save(f'{currTimeStr}-report.xlsx')
wb = openpyxl.load_workbook(f'{currTimeStr}-report.xlsx')
if flag in wb.sheetnames:
    ws_Nexus = wb[flag]
else:
    ws_Nexus = wb.create_sheet(title=flag)

ws_Nexus['A1'] = "IP地址"
ws_Nexus['B1'] = '设备型号'
ws_Nexus['C1'] = '设备版本'
ws_Nexus['D1'] = '设备运行时间'
ws_Nexus['E1'] = 'CPU使用率'
ws_Nexus['F1'] = '内存使用率'
ws_Nexus['G1'] = 'Fex状态'
ws_Nexus['H1'] = '风扇状态'
ws_Nexus['I1'] = '电源状态'
ws_Nexus['J1'] = '设备温度状态'
ws_Nexus['K1'] = '百兆接口'
ws_Nexus['L1'] = 'errdisable接口'
# ws_ios['M1'] = '根桥地址'

wsline = 1
for file in files:
    wsline += 1
    valuesDict = processHuaWeiFile('172.30.0.1-2018-07-31.txt', CommandString)
    ws_Nexus[f'A{wsline}'] = valuesDict['IP']
    ws_Nexus[f'B{wsline}'] = valuesDict['Type']
    ws_Nexus[f'C{wsline}'] = valuesDict['Version']
    ws_Nexus[f'D{wsline}'] = valuesDict['UpTime']
    ws_Nexus[f'E{wsline}'] = valuesDict['CpuUsage']
    ws_Nexus[f'F{wsline}'] = valuesDict['MemoryUsage']
    ws_Nexus[f'G{wsline}'] = valuesDict['Fex']
    ws_Nexus[f'H{wsline}'] = valuesDict['Fan']
    ws_Nexus[f'I{wsline}'] = valuesDict['Power']
    ws_Nexus[f'J{wsline}'] = valuesDict['Temperature']
    ws_Nexus[f'K{wsline}'] = valuesDict['100M']
    ws_Nexus[f'L{wsline}'] = valuesDict['ErrDisable']
#     # ws_ios[f'M{wsline}'] = valuesDict['treeRootMacadd']
wb.save(f'{currTimeStr}-report.xlsx')
