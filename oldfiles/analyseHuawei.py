#!/usr/bin/env python  
#coding:utf-8  
"""
@author:leo
@license: Apache Licence 
@file: analyseConfig.py.py 
@time: 2018/08/01
@contact: leiaoo@163.com
"""
import os, datetime, re, json
import openpyxl

ipStrRe = re.compile(r'(\d{1,3}(\.)?){1,4}')
cpuStrRe = re.compile(r'\d{0,2}\.?\d{0,2}%')
# currTimeStr = datetime.datetime.now().strftime('%Y-%m-%d')
currTimeStr = '2018-08-09'
baseDir = 'device'
flag = 'HuaWei'
files = os.listdir(os.path.join(baseDir, flag, currTimeStr))
with open('conf_files/HuaweiCommand.txt', 'r') as f:
    CommandString = json.load(f)
valuesList = {}

def processList(valuesList, startIndex, endIndex):
    failedFlag = 'Normal'
    for line in valuesList[startIndex:endIndex]:
        if '-' * 10 in line:
            continue
        elif line == '\n':
            continue
        elif 'normal' in line.lower():
            continue
        elif 'yes' in line.lower():
            continue
        elif 'ac' in line.lower() and 'supply' in line.lower():
            continue
        else:
            failedFlag = 'Failed'
            break
    return failedFlag

def processHuaWeiFile(file, CommandString, ):
    valuesDict = {
        'IP': 'None', #device file IP
        'flag': flag, #device type
        'Version': 'None', #device version
        'Type' : 'None',
        'Bgp_Peer': 'None', #bgp peer information
        'Ospf_Peer': 'None', #ospf peer information
        'Ospf_brief': 'None', #ospf information
        'UpTime': 'None',  #device up time
        'Health': 'None',  #device health information
        'Fan': 'None',  #device fan status
        'Power': 'None', #device power status
        'Temperature': 'None', #device temperature status
        'DeviceStatus': 'None', #device status
        'CpuUsage': 'None', #cpu usage
        'MemoryUsage': 'None', # memory usage
        'InterfaceStatus': {  #interface information
            'InterfaceNum': 'None', #interface Number
            'PHY': 'None',  #interface physics status
            'Protocol': 'None', #interface protocol status
            'InputErrors': 'None', #interface input error
            'OutputErrors': 'None', #interface output error
            'InUti': 'None', #interface input utilities
            'OutUti': 'None', #interface output utilities
            'Speed': 'None', #interface speed
            'Duplex': 'None', #interface duplex module
        },
    }
    commandIndex = {}
    valuesDict['IP'] = ipStrRe.search(file).group()
    file = os.path.join(baseDir, flag, currTimeStr, file)

    with open(file, 'r') as f:
        fileReadLines = f.readlines()

    for i in range(len(CommandString)):
        commandIndex[CommandString[i]] = fileReadLines.index(CommandString[i] + '\n')

    valuesList['bgp'] = fileReadLines[commandIndex['display bgp peer']:commandIndex['display ospf peer']]
    valuesList['ospf'] = fileReadLines[commandIndex['display ospf peer']:commandIndex['display version']]
    valuesList['version'] = fileReadLines[commandIndex['display version']:commandIndex['display health']]
    valuesList['health'] = fileReadLines[commandIndex['display health']:commandIndex['display fan']]
    valuesList['fan'] = fileReadLines[commandIndex['display fan']:commandIndex['display power']]
    valuesList['power'] = fileReadLines[commandIndex['display power']:commandIndex['display temperature']]
    valuesList['temperature'] = fileReadLines[commandIndex['display temperature']:commandIndex['display temperature all']]
    valuesList['temperatureAll'] = fileReadLines[commandIndex['display temperature all']:commandIndex['display device']]
    valuesList['device'] = fileReadLines[commandIndex['display device']:commandIndex['display cpu-usage']]
    valuesList['cpu'] = fileReadLines[commandIndex['display cpu-usage']:commandIndex['display memory-usage']]
    valuesList['memory'] = fileReadLines[commandIndex['display memory-usage']:commandIndex['display interface']]
    valuesList['disInt'] = fileReadLines[commandIndex['display interface']:commandIndex['display ip interface']]
    valuesList['disIpInt'] = fileReadLines[commandIndex['display ip interface']:commandIndex['display interface brief']]
    valuesList['disIntBri'] = fileReadLines[commandIndex['display interface brief']:commandIndex['display counters']]
    valuesList['Counters'] = fileReadLines[commandIndex['display counters']:commandIndex['display counters error']]
    valuesList['CounErr'] = fileReadLines[commandIndex['display counters error']:commandIndex['display counters rate']]
    valuesList['CounRate'] = fileReadLines[commandIndex['display counters rate']:commandIndex['display alarm statistics']]
    valuesList['alarmSta'] = fileReadLines[commandIndex['display alarm statistics']:commandIndex['display alarm information']]
    valuesList['alarmInfo'] = fileReadLines[commandIndex['display alarm information']:commandIndex['display trapbuffer']]
    valuesList['trapbuffer'] = fileReadLines[commandIndex['display trapbuffer']:commandIndex['display logbuffer']]
    valuesList['logbuffer'] = fileReadLines[commandIndex['display logbuffer']:]

    version = [line.split(',')[1].strip('\n') for line in valuesList['version'] if 'software, version'
                             in line.lower()][0].split(' ')
    valuesDict['Version'] = version[2]
    valuesDict['Type'] = version[3].strip('(')
    valuesDict['UpTime'] = str(int([line.split(':')[1].strip('\n') for line in valuesList['version'] if ': uptime'
                             in line.lower()][0].strip(' ').split(' ')[2]) * 7)
    valuesDict['MemoryUsage'] = [line.split(':')[1].strip('\n') for line in valuesList['memory'] if 'memory Using'.lower()
                             in line.lower()][0].strip(' ')
    cpuStr = [line for line in valuesList['cpu'] if 'CPU Usage'.lower()
                         in line.lower() and 'Max'.lower() in line.lower()][0].strip('\n').strip(' ')
    valuesDict['CpuUsage'] = cpuStrRe.search(cpuStr).group()
    for key, value in valuesList.items():
        Flag = 'Normal'
        if len(valuesList[key]) <= 3:
            continue
        if key == 'fan':
            valuesDict['Fan'] = 'Normal' if 'normal' in valuesList[key][4].lower() or 'yes' in \
                                            valuesList[key][4].lower() else 'Failed'
        elif key == 'health':
            startIndex, endIndex = 4, 9
            valuesDict['Health'] = processList(valuesList[key], startIndex, endIndex)
        elif 'power' in key:
            startIndex, endIndex = 4, -1
            valuesDict['Power'] = processList(valuesList[key], startIndex, endIndex)
        elif 'temperature' in key:
            startIndex, endIndex = 4, -1
            valuesDict['Temperature'] = processList(valuesList[key], startIndex, endIndex)
        elif 'device' in key:
            startIndex, endIndex = 4, -1
            valuesDict['DeviceStatus'] = processList(valuesList[key], startIndex, endIndex)
    return valuesDict


if not os.path.exists(f'{currTimeStr}-report.xlsx'):
    wb = openpyxl.Workbook()
    wb.save(f'{currTimeStr}-report.xlsx')
wb = openpyxl.load_workbook(f'{currTimeStr}-report.xlsx')

if 'Sheet' in wb.sheetnames:
    ws_huawei = wb['Sheet']
    ws_huawei.title = flag
elif flag in wb.sheetnames:
    ws_huawei = wb[flag]
else:
    ws_huawei = wb.create_sheet(title=flag)

ws_huawei = wb[flag]
ws_huawei['A1'] = "IP地址"
ws_huawei['B1'] = '设备版本'
ws_huawei['C1'] = '设备型号'
ws_huawei['D1'] = '设备运行时间'
ws_huawei['E1'] = '内存使用率'
ws_huawei['F1'] = 'CPU使用率'
ws_huawei['G1'] = '风扇状态'
ws_huawei['H1'] = '设备健康状态'
ws_huawei['I1'] = '电源状态'
ws_huawei['J1'] = '设备温度状态'
ws_huawei['K1'] = '设备状态'
wsline = 1
for file in files:
    wsline += 1
    valuesDict = processHuaWeiFile(file, CommandString)
    ws_huawei[f'A{wsline}'] = valuesDict['IP']
    ws_huawei[f'B{wsline}'] = valuesDict['Version']
    ws_huawei[f'C{wsline}'] = valuesDict['Type']
    ws_huawei[f'D{wsline}'] = valuesDict['UpTime'] + '天'
    ws_huawei[f'E{wsline}'] = valuesDict['MemoryUsage']
    ws_huawei[f'F{wsline}'] = valuesDict['CpuUsage']
    ws_huawei[f'G{wsline}'] = valuesDict['Fan']
    ws_huawei[f'H{wsline}'] = valuesDict['Health']
    ws_huawei[f'I{wsline}'] = valuesDict['Power']
    ws_huawei[f'J{wsline}'] = valuesDict['Temperature']
    ws_huawei[f'K{wsline}'] = valuesDict['DeviceStatus']
wb.save(f'{currTimeStr}-report.xlsx')
